import openpyxl

file_path = "TestData.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Correcting headers based on professional SDDF standards
sheet.cell(row=1, column=1, value="TestID")
sheet.cell(row=1, column=2, value="username")
sheet.cell(row=1, column=3, value="password")
sheet.cell(row=1, column=4, value="expected_error")
sheet.cell(row=1, column=5, value="expected_title")
sheet.cell(row=1, column=6, value="Result")        # Dynamic write target
sheet.cell(row=1, column=7, value="Execution_Time") # Dynamic write target

wb.save(file_path)
print("Excel Headers Restored to SDDF Standard: [TestID, username, password, expected_error, expected_title, Result, Execution_Time]")
